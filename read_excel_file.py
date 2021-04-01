import openpyxl

path = '/home/theo/Downloads/Test/data3.xlsx'

workbook = openpyxl.load_workbook(path)
# sheet = workbook.active
sheet = workbook.get_sheet_by_name('Sheet1')

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)


for row in range(1, rows+1):
    for col in range(1, cols+1):
        data = sheet.cell(row=row, column=col).value
        print(data, end='     \t')
    print()